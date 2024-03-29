import datetime
import json
import os
import openpyxl
import time

from src.path_info import database_folder, export_folder
from openpyxl.workbook import Workbook
from src.crawler import BannerCrawler


class User:
    def __init__(self, uid):
        self.UID = uid
        self.path = f'{database_folder}\\user\\{self.UID}.json'
        self._data = None
        self.CharacterBanner = []
        self.WeaponBanner = []
        self.NormalBanner = []
        self.read()
        self.api_url = None

    def set_character_banner(self, history_data):
        self.CharacterBanner = self.add_history_data(self.CharacterBanner, history_data)
        return self

    def set_api_url(self, api_url):
        self.api_url = api_url
        print(f'set api {api_url}')
        return self

    def set_weapon_banner(self, history_data):
        self.WeaponBanner = self.add_history_data(self.WeaponBanner, history_data)
        return self

    def set_normal_banner(self, history_data):
        self.NormalBanner = self.add_history_data(self.NormalBanner, history_data)
        return self

    def read(self):
        if os.path.isfile(self.path):
            with open(self.path, 'r', encoding='utf-8') as u:
                self._data = json.loads(u.read())
                self.CharacterBanner = self._data['wish_history']['character_banner']
                self.WeaponBanner = self._data['wish_history']['weapon_banner']
                self.NormalBanner = self._data['wish_history']['normal_banner']
                try:
                    self.api_url = self._data['api_url']
                except KeyError:
                    self.api_url = ''
        else:
            self.save()
            self.read()
        return self

    def __dict__(self):
        user_dict = {
            'uid': self.UID,
            'wish_history': {
                'character_banner': self.CharacterBanner,
                'weapon_banner': self.WeaponBanner,
                'normal_banner': self.NormalBanner,
            },
            'api_url': self.api_url,
            'time': datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S").__str__()
        }
        return user_dict

    def save(self):
        with open(self.path, 'w', encoding='utf-8', errors='ignore') as u:
            json.dump(self.__dict__(), u, ensure_ascii=False)
        return self

    @staticmethod
    def add_history_data(current_data: list, last_data: list):
        if current_data is None:
            current_data = []
        if len(last_data) == 0:
            return current_data
        last_data_timestamp = last_data[-1]['id']
        j = 0
        for i in range(len(current_data)):
            if current_data[i]['id'] < last_data_timestamp:
                j = i
                break
        new_data = last_data
        new_data.extend(current_data[j:])
        return new_data

    def export_xlsx(self, output_file=None, ignore_3_star=False):
        if output_file is None:
            output_file = export_folder + f'\\{self.UID}_{int(time.time())}.xlsx'
        workbook = None
        if os.path.isfile(output_file):
            workbook = openpyxl.load_workbook(output_file)
        else:
            workbook = Workbook()
        if 'CharacterBanner' not in workbook.sheetnames:
            workbook.create_sheet('CharacterBanner')
        if 'WeaponBanner' not in workbook.sheetnames:
            workbook.create_sheet('WeaponBanner')
        if 'NormalBanner' not in workbook.sheetnames:
            workbook.create_sheet('NormalBanner')

        for value in tuple(BannerCrawler.history_to_array(self.CharacterBanner, ignore_3_star=ignore_3_star)):
            workbook['CharacterBanner'].append(value)
        for value in tuple(BannerCrawler.history_to_array(self.WeaponBanner, ignore_3_star=ignore_3_star)):
            workbook['WeaponBanner'].append(value)
        for value in tuple(BannerCrawler.history_to_array(self.NormalBanner, ignore_3_star=ignore_3_star)):
            workbook['NormalBanner'].append(value)

        workbook.save(output_file)
        return output_file

    def get_last_character_banner_id(self):
        if self.CharacterBanner is None or len(self.CharacterBanner) == 0:
            return 0
        else:
            return self.CharacterBanner[0]['id']

    def get_last_weapon_banner_id(self):
        if self.WeaponBanner is None or len(self.WeaponBanner) == 0:
            return 0
        else:
            return self.WeaponBanner[0]['id']

    def get_last_normal_banner_id(self):
        if self.NormalBanner is None or len(self.NormalBanner) == 0:
            return 0
        else:
            return self.NormalBanner[0]['id']
