import os

from openpyxl import Workbook
from src import User, path_info
from src.util import get_file_names


def export_to_paimon_moe_xlsx(user: User, output_path: str = None, __sort='desc'):
    if output_path is None:
        output_path = path_info.export_folder + f'\\{user.UID}_paimon_moe.xlsx'
    workbook = _paimon_moe_xlsx_workbook()

    if __sort is 'desc':
        character_his = user.CharacterBanner.__reversed__()
        weapon_his = user.WeaponBanner.__reversed__()
        normal_his = user.NormalBanner.__reversed__()
    else:
        character_his = user.CharacterBanner
        weapon_his = user.WeaponBanner
        normal_his = user.NormalBanner

    for h in character_his:
        workbook['Character Event'].append(
            (h['item_type'], h['name'], h['time'], h['rank_type'])
        )
    for h in weapon_his:
        workbook['Weapon Event'].append(
            (h['item_type'], h['name'], h['time'], h['rank_type'])
        )
    for h in normal_his:
        workbook['Standard'].append(
            (h['item_type'], h['name'], h['time'], h['rank_type'])
        )

    workbook.save(output_path)
    return output_path


def _paimon_moe_xlsx_workbook():
    workbook = Workbook()
    workbook.create_sheet('Information')
    workbook.create_sheet('Character Event')
    workbook.create_sheet('Weapon Event')
    workbook.create_sheet('Standard')
    workbook.create_sheet("Beginners' Wish")

    workbook['Information'].append(('Paimon.moe Wish History Export',))
    workbook['Information'].append(('Version', '3'))
    workbook['Information'].append(('Export Date', '3'))
    header = ('Type', 'Name', 'Time', '⭐', 'Pity', '#Roll', 'Group', 'Banner', 'Part')
    workbook['Character Event'].append(header)
    workbook['Weapon Event'].append(header)
    workbook['Standard'].append(header)
    return workbook


def delete_exported_xlsx(uid):
    path = path_info.export_folder
    files_name = get_file_names(path)

    if len(files_name) != 0:
        for fn in files_name:
            if fn.startswith(str(uid)):
                os.remove(f'{path}\\{fn}')
    return True
